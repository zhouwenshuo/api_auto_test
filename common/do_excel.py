# -*- coding: utf-8 -*-
'''
@time: 2019/4/15 14:21
@author: beijing-Mr.Right-15
@contact: 348533713@qq.com
@file: do_excel.py
@desc:
        ┏┓　　　┏┓+ +
　　　┏┛┻━━━┛┻┓ + +
　　　┃　　　　　　　┃ 　
　　　┃　　　━　　　┃ ++ + + +
　　 ████━████ ┃+
　　　┃　　　　　　　┃ +
　　　┃　　　┻　　　┃
　　　┃　　　　　　　┃ + +
　　　┗━┓　　　┏━┛
　　　　　┃　　　┃　　　　　　　　　　　
　　　　　┃　　　┃ + + + +
　　　　　┃　　　┃　　　　Codes are far away from bugs with the animal protecting　　　
　　　　　┃　　　┃ + 　　　　神兽保佑,代码无bug　　
　　　　　┃　　　┃
　　　　　┃　　　┃　　+　　　　　　　　　
　　　　　┃　 　　┗━━━┓ + +
　　　　　┃ 　　　　　　　┣┓
　　　　　┃ 　　　　　　　┏┛
　　　　　┗┓┓┏━┳┓┏┛ + + + +
　　　　　　┃┫┫　┃┫┫
　　　　　　┗┻┛　┗┻┛+ + + +
excel操作，用例操作
'''

from openpyxl import load_workbook
from common import contants


class DoExcel:

    # def __init__(self,sheet_name,file_name=contants.case_file): #第一种：一个excel多个sheet页
    #     self.filename=file_name
    #     self.wb = load_workbook(file_name)
    #     self.sheet = self.wb[sheet_name]

    def __init__(self, file_name, file_path=contants.case_file):  # 第二种：多个excel，一个excel是一个接口
        self.filename = file_name
        self.filepath = file_path + file_name + '.xlsx'
        self.wb = load_workbook(self.filepath)
        self.sheet = self.wb[file_name]

    def get_data_from_excel(self):
        '''从excel获取测试数据'''
        datas = []
        for row in range(2, self.sheet.max_row + 1):
            row_data = []
            for column in range(1, 7):
                row_data.append(self.sheet.cell(row, column).value)
            datas.append(row_data)
        self.wb.close()
        return datas

    # def write_result(self,row,actual=None,result=None):  #第一种
    #     if actual==None:
    #         self.sheet.cell(row, 8).value = result
    #     elif result==None:
    #         self.sheet.cell(row,7).value=actual
    #     self.wb.save(filename=self.filename)
    #     self.wb.close()

    def write_result(self, row, actual=None, result=None):  # 第二种
        if actual == None:
            self.sheet.cell(row, 8).value = result
        elif result == None:
            self.sheet.cell(row, 7).value = actual
        self.wb.save(filename=self.filepath)
        self.wb.close()

    def close(self):
        self.wb.close()


if __name__ == '__main__':
    obj = DoExcel(file_name='TestLogin')
    obj.get_data_from_excel()
